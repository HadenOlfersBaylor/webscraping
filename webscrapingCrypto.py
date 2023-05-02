from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

webpage = 'https://www.coinlore.com/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

#  The report should display the name of the currency, the symbol (if applicable), 
#  the current price and % change in the last 24 hrs and corresponding price (based on % change)

wb = xl.Workbook()

ws = wb.active

ws.title = 'Crypto Report'

ws['A1'] = 'No.'
ws['B1'] = 'Crypto Title'
ws['C1'] = 'Symbol'
ws['D1'] = 'Price'
ws['E1'] = '24hr % Change'
ws['F1'] = 'Corresponding Price'

crypto_rows = soup.findAll('tr')

for x in range(1,6):
    td = crypto_rows[x].findAll('td')
    sym = crypto_rows[x].findAll('small')
    no = td[1].text
    title = td[3].text
    symbol = sym[0].text
    price = float(td[4].text.replace(",","").replace("$",""))
    gross = float(td[6].text.replace(",","").replace("%",""))

    corr_price = price * (1+gross)

    print(f"{no}, {title}, {symbol}, {price}, {gross}, {corr_price}")

    ws['A' + str(x+1)] = no
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = symbol
    ws['D' + str(x+1)] = price
    ws['E' + str(x+1)] = str(gross) + '%'
    ws['F' + str(x+1)] = corr_price

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws['D:D']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['F:F']:
    cell.number_format = u'"$ "#,##0.00'
    

wb.save('CryptoReport.xlsx')